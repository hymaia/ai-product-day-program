import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('/Users/elsamargier/Desktop/AI product day/[AI Product Day 26] ProgrammeV2.xlsx')
for sheet_name in wb.sheetnames:
    print(f'\n=== {sheet_name} ===')
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.type != 'none':
                color = fill.fgColor.rgb if fill.fgColor.type == 'rgb' else str(fill.fgColor.value)
                if cell.value and color not in ('00000000', 'FFFFFFFF', '00FFFFFF', 'FF000000'):
                    print(f'  [{cell.coordinate}] color={color} | value={str(cell.value)[:80]}')
